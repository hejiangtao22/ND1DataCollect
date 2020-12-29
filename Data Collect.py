import openpyxl
import pandas as pd
import datetime
import os
from datetime import datetime

file_folder_DNA = r'C:\Python\ND1 Data Collect\DS_DNA'
file_folder_Protein = r'C:\Python\ND1 Data Collect\Protein'
file_folder_UV_Vis = r'C:\Python\ND1 Data Collect\UV-Vis'
file_folder = r'C:\Python\ND1 Data Collect'
file_data_folder = r'C:\Python\ND1 Data Collect\Data'
file_summary = openpyxl.load_workbook(r'C:\Python\ND1 Data Collect\template.xlsx')
sheet_file_summary = file_summary.get_sheet_by_name('Summary')
list_file_data = os.listdir(file_data_folder)
list_file_DNA = os.listdir(file_folder_DNA)
list_file_Protein = os.listdir(file_folder_Protein)
list_file_UV_Vis = os.listdir(file_folder_UV_Vis)
print('list file UV_Vis:', list_file_UV_Vis)

global current_row
current_row = 1
for f in list_file_data:
    application = pd.read_csv(f, sep='\s+', skiprows=0, nrows=1, header=None, encoding = 'utf-16')[1][0]    # read application name
    print(application)

    if application == 'dsDNA':
        serial_number = pd.read_csv(f, sep='\s+', skiprows=1, nrows=1, header=None, encoding = 'utf-16')[2][0]  # read serial number
        print(serial_number)
        column_names = pd.read_csv(f, sep='\s+', skiprows=2, nrows=1, header=None, encoding = 'utf-16').values.tolist() # read column names
        print(column_names)
        df_data = pd.read_csv(f, sep='\s+', skiprows=3, header=None, encoding = 'utf-16')   # read all data
        print(df_data)
        i = 0
        for i in range(len(df_data)):
            current_row += 1
            sheet_file_summary.cell(row = current_row, column = 1).value = serial_number  # add serial number
            sheet_file_summary.cell(row = current_row, column=2).value = application  # add application name
            sheet_file_summary.cell(row = current_row, column= 4).value = df_data[4][i] + '_' + df_data[5][i]   # add Sample ID
            sheet_file_summary.cell(row = current_row, column=5).value = df_data[7][i]      # add concentration
            sheet_file_summary.cell(row = current_row, column=7).value = df_data[10][i]     # add A260 Abs. to current Abs.
            sheet_file_summary.cell(row = current_row, column=16).value = str(df_data[12][i]) + '_' + str(df_data[13][i]) + '_' + str(df_data[14][i])    # add pathlength

    if application == 'UV-Vis':
        serial_number = pd.read_csv(f, sep='\s+', skiprows=1, nrows=1, header=None, encoding='utf-16')[2][0]  # read serial number
        print(serial_number)
        column_names = pd.read_csv(f, sep='\s+', skiprows=2, nrows=1, header=None, encoding='utf-16').values.tolist()  # read column names
        print(column_names)
        df_data = pd.read_csv(f, sep='\s+', skiprows=3, header=None, encoding='utf-16')  # read all data
        print(df_data)
        i = 0
        for i in range(len(df_data)):
            current_row += 1
            sheet_file_summary.cell(row=current_row, column=1).value = serial_number  # add serial number
            sheet_file_summary.cell(row=current_row, column=2).value = application  # add application name
            sheet_file_summary.cell(row=current_row, column=4).value = str(df_data[3][i]) + '_' + str(df_data[4][i])  # add Sample ID
            sheet_file_summary.cell(row=current_row, column=12).value = df_data[15][i]   # add Abs of UV260nm
            sheet_file_summary.cell(row=current_row, column=13).value = df_data[16][i]  # add Abs of UV340nm
            sheet_file_summary.cell(row=current_row, column=14).value = df_data[17][i]  # add Abs of UV416nm
            sheet_file_summary.cell(row=current_row, column=15).value = df_data[18][i]  # add Abs of UV627nm
            sheet_file_summary.cell(row=current_row, column=16).value = str(df_data[12][i]) + '_' + str(df_data[13][i]) + '_' + str(df_data[14][i])  # add pathlength

    if application == 'ProteinA205':
        serial_number = pd.read_csv(f, sep='\s+', skiprows=1, nrows=1, header=None, encoding='utf-16')[2][0]  # read serial number
        print(serial_number)
        column_names = pd.read_csv(f, sep='\s+', skiprows=2, nrows=1, header=None, encoding='utf-16').values.tolist()  # read column names
        print(column_names)
        df_data = pd.read_csv(f, sep='\s+', skiprows=3, header=None, encoding='utf-16')  # read all data
        print(df_data)
        i = 0
        for i in range(len(df_data)):
            current_row += 1
            sheet_file_summary.cell(row=current_row, column=1).value = serial_number  # add serial number
            sheet_file_summary.cell(row=current_row, column=2).value = application  # add application name
            sheet_file_summary.cell(row=current_row, column=4).value = str(df_data[3][i]) + '_' + str(df_data[4][i])  # add Sample ID
            sheet_file_summary.cell(row=current_row, column=5).value = df_data[6][i]  # add concentration
            sheet_file_summary.cell(row=current_row, column=7).value = df_data[7][i]  # add Abs of A205nm to current Abs
            sheet_file_summary.cell(row=current_row, column=16).value = str(df_data[10][i]) + '_' + str(df_data[11][i]) + '_' + str(df_data[12][i])  # add pathlength
            sheet_file_summary.cell(row=current_row, column=17).value = df_data[9][i]  # add method

file_summary.save(file_folder + '\Data_Summary_' + datetime.now().strftime('%Y-%m-%d') + datetime.now().strftime('-%H-%M-%S') + '.xlsx')




